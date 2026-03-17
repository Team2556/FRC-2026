"""
Generate the SysId Characterization Job Aide & Input Form workbook (.xlsx).

Run:  python docs/generate_sysid_workbook.py
Output: docs/SysId_Characterization_Workbook.xlsx
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ── Styles ───────────────────────────────────────────────────────────
HEADER_FONT = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
SECTION_FONT = Font(name="Calibri", bold=True, size=11, color="1F4E79")
SECTION_FILL = PatternFill("solid", fgColor="D6E4F0")
TITLE_FONT = Font(name="Calibri", bold=True, size=16, color="1F4E79")
SUBTITLE_FONT = Font(name="Calibri", bold=False, size=10, color="555555")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")  # light yellow
PASS_FILL = PatternFill("solid", fgColor="C6EFCE")
FAIL_FILL = PatternFill("solid", fgColor="FFC7CE")
WARN_FILL = PatternFill("solid", fgColor="FFE699")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)


def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def style_section_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = SECTION_FONT
        cell.fill = SECTION_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def style_input_cell(ws, row, col):
    cell = ws.cell(row=row, column=col)
    cell.fill = INPUT_FILL
    cell.border = THIN_BORDER
    cell.alignment = CENTER
    return cell


def apply_all_borders(ws, min_row, max_row, max_col):
    for r in range(min_row, max_row + 1):
        for c in range(1, max_col + 1):
            ws.cell(row=r, column=c).border = THIN_BORDER


# =====================================================================
# SHEET 1: Cover / Pre-Flight Checklist
# =====================================================================
def build_cover_sheet(wb):
    ws = wb.active
    ws.title = "1 - Pre-Flight"
    ws.sheet_properties.tabColor = "1F4E79"

    # Title
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = "Team 2556 - SysId Drivetrain Characterization"
    c.font = TITLE_FONT
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = "Job Aide & Data Collection Workbook"
    c.font = SUBTITLE_FONT
    c.alignment = Alignment(horizontal="center")

    # Session info
    row = 4
    info_labels = [
        "Date:",
        "Location:",
        "Robot Name / ID:",
        "Battery Voltage (start):",
        "Firmware Versions Checked:",
        "Operator Name:",
        "Driver Station Laptop:",
    ]
    for label in info_labels:
        ws.cell(row=row, column=1, value=label).font = SECTION_FONT
        ws.cell(row=row, column=1).alignment = LEFT_WRAP
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        style_input_cell(ws, row, 2)
        row += 1

    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws.cell(row=row, column=1, value="PRE-FLIGHT CHECKLIST").font = SECTION_FONT
    ws.cell(row=row, column=1).fill = SECTION_FILL
    ws.cell(row=row, column=1).alignment = CENTER
    row += 1

    checklist = [
        ("Robot on blocks / wheels off ground", "REQUIRED for translation & steer tests to prevent runaway"),
        ("Battery fully charged (> 12.5V)", "Low voltage skews SysId results — replace if < 12.5V"),
        ("All CAN devices healthy (no faults)", "Check Tuner X / Phoenix Diagnostics before starting"),
        ("Code deployed (feature/sysid-drivetrain)", "Verify correct branch is deployed to roboRIO"),
        ("Controller 3 connected (USB port 2)", "Dedicated SysId controller — verify in Driver Station USB tab"),
        ("Driver Station in Test mode", "SysId runs in Test mode — SignalLogger starts automatically"),
        ("Sufficient free space on roboRIO", "HOOT logs need ~50-100 MB — check /home/lvuser/logs/"),
        ("All team members clear of robot", "Robot WILL move during characterization — safety first"),
        ("E-stop accessible", "Keep e-stop or Enter key ready at all times"),
    ]

    headers = ["#", "Item", "Details / Why", "Pass?"]
    col_widths = [4, 42, 55, 8]
    for i, (h, w) in enumerate(zip(headers, col_widths), 1):
        ws.cell(row=row, column=i, value=h)
        ws.column_dimensions[get_column_letter(i)].width = w
    style_header_row(ws, row, len(headers))
    row += 1

    for idx, (item, detail) in enumerate(checklist, 1):
        ws.cell(row=row, column=1, value=idx).alignment = CENTER
        ws.cell(row=row, column=2, value=item).alignment = LEFT_WRAP
        ws.cell(row=row, column=3, value=detail).alignment = LEFT_WRAP
        style_input_cell(ws, row, 4)
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=2).border = THIN_BORDER
        ws.cell(row=row, column=3).border = THIN_BORDER
        ws.row_dimensions[row].height = 30
        row += 1

    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.cell(row=row, column=1).value = (
        "IMPORTANT: Do NOT proceed until ALL checklist items pass."
    )
    ws.cell(row=row, column=1).font = Font(bold=True, color="CC0000", size=11)

    # Column widths
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12


# =====================================================================
# SHEET 2: Controller Map & Procedure
# =====================================================================
def build_procedure_sheet(wb):
    ws = wb.create_sheet("2 - Procedure")
    ws.sheet_properties.tabColor = "2E75B6"

    ws.merge_cells("A1:G1")
    ws["A1"].value = "SysId Test Procedure & Controller Map"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center")

    # Controller map
    row = 3
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws.cell(row=row, column=1, value="CONTROLLER 3 (Port 2) — Button Map").font = SECTION_FONT
    ws.cell(row=row, column=1).fill = SECTION_FILL
    row += 1

    btn_headers = ["Button", "Action", "What It Does", "Hold / Press", "Duration"]
    for i, h in enumerate(btn_headers, 1):
        ws.cell(row=row, column=i, value=h)
    style_header_row(ws, row, len(btn_headers))
    row += 1

    buttons = [
        ("A", "Quasistatic Forward", "Slowly ramps voltage forward (1 V/s)", "HOLD", "~10s"),
        ("B", "Quasistatic Reverse", "Slowly ramps voltage reverse (1 V/s)", "HOLD", "~10s"),
        ("X", "Dynamic Forward", "Steps to full test voltage forward", "HOLD", "~3-5s"),
        ("Y", "Dynamic Reverse", "Steps to full test voltage reverse", "HOLD", "~3-5s"),
        ("POV Left", "Select: Translation", "Switches to drive motor characterization", "TAP", "Instant"),
        ("POV Up", "Select: Steer", "Switches to steer/azimuth motor characterization", "TAP", "Instant"),
        ("POV Right", "Select: Rotation", "Switches to heading/rotation characterization", "TAP", "Instant"),
    ]

    for btn, action, desc, hold, dur in buttons:
        ws.cell(row=row, column=1, value=btn).alignment = CENTER
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2, value=action).alignment = LEFT_WRAP
        ws.cell(row=row, column=3, value=desc).alignment = LEFT_WRAP
        ws.cell(row=row, column=4, value=hold).alignment = CENTER
        ws.cell(row=row, column=5, value=dur).alignment = CENTER
        for c in range(1, 6):
            ws.cell(row=row, column=c).border = THIN_BORDER
        row += 1

    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws.cell(row=row, column=1, value="STEP-BY-STEP PROCEDURE").font = SECTION_FONT
    ws.cell(row=row, column=1).fill = SECTION_FILL
    row += 1

    steps = [
        ("1", "SETUP", "Complete ALL Pre-Flight checklist items. Robot on blocks, battery charged, code deployed."),
        ("2", "ENABLE TEST MODE", "In Driver Station, select Test mode and Enable. SignalLogger starts automatically. Confirm no errors in DS console."),
        ("3", "SELECT ROUTINE", "Press POV Left on Controller 3 to select Translation (default). Confirm on DS console or SmartDashboard."),
        ("4", "QUASISTATIC FWD", "Hold A — robot wheels spin slowly with increasing speed. Hold for full ~10 seconds, then release. Record battery voltage."),
        ("5", "PAUSE", "Wait 2-3 seconds between each test for data separation in the log."),
        ("6", "QUASISTATIC REV", "Hold B — same as step 4 but in reverse direction. Hold ~10 seconds, release."),
        ("7", "PAUSE", "Wait 2-3 seconds."),
        ("8", "DYNAMIC FWD", "Hold X — wheels instantly jump to test voltage (4V for translation). Hold 3-5 seconds, release."),
        ("9", "PAUSE", "Wait 2-3 seconds."),
        ("10", "DYNAMIC REV", "Hold Y — same as step 8 but reverse. Hold 3-5 seconds, release."),
        ("11", "CHECK BATTERY", "Read battery voltage on DS. If < 12.0V, replace battery before continuing. Record voltage."),
        ("12", "SWITCH TO STEER", "Press POV Up. Repeat steps 4-11 for Steer routine. Note: steer uses 7V dynamic step."),
        ("13", "SWITCH TO ROTATION", "Press POV Right. Repeat steps 4-11 for Rotation routine. Note: uses angular rate, not voltage."),
        ("14", "DISABLE", "Disable the robot. SignalLogger stops automatically and writes the HOOT file."),
        ("15", "RETRIEVE LOG", "Connect to roboRIO via USB/SSH. Copy HOOT file from /home/lvuser/logs/ to your laptop."),
        ("16", "ANALYZE", "Open the HOOT file in the WPILib SysId tool. Run analysis for each routine. Record gains on the Results sheet."),
    ]

    step_headers = ["Step", "Phase", "Instructions"]
    for i, h in enumerate(step_headers, 1):
        ws.cell(row=row, column=i, value=h)
    style_header_row(ws, row, 3)
    row += 1

    for step, phase, instr in steps:
        ws.cell(row=row, column=1, value=step).alignment = CENTER
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2, value=phase).alignment = CENTER
        ws.cell(row=row, column=2).font = Font(bold=True)
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=5)
        ws.cell(row=row, column=3, value=instr).alignment = LEFT_WRAP
        for c in range(1, 6):
            ws.cell(row=row, column=c).border = THIN_BORDER
        ws.row_dimensions[row].height = 35
        row += 1

    # Voltage reference
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws.cell(row=row, column=1, value="TEST VOLTAGE REFERENCE").font = SECTION_FONT
    ws.cell(row=row, column=1).fill = SECTION_FILL
    row += 1

    volt_headers = ["Routine", "Quasistatic Ramp", "Dynamic Step Voltage", "Timeout", "Notes"]
    for i, h in enumerate(volt_headers, 1):
        ws.cell(row=row, column=i, value=h)
    style_header_row(ws, row, 5)
    row += 1

    volt_data = [
        ("Translation", "1 V/s (default)", "4V", "10s", "Lower voltage prevents brownout under load"),
        ("Steer", "1 V/s (default)", "7V", "10s", "Azimuth motors need less current"),
        ("Rotation", "pi/6 rad/s^2", "7V (as rad/s)", "10s", "Units are angular — SysId interprets accordingly"),
    ]
    for routine, ramp, dyn, timeout, notes in volt_data:
        ws.cell(row=row, column=1, value=routine).alignment = CENTER
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2, value=ramp).alignment = CENTER
        ws.cell(row=row, column=3, value=dyn).alignment = CENTER
        ws.cell(row=row, column=4, value=timeout).alignment = CENTER
        ws.cell(row=row, column=5, value=notes).alignment = LEFT_WRAP
        for c in range(1, 6):
            ws.cell(row=row, column=c).border = THIN_BORDER
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 48
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14


# =====================================================================
# SHEET 3: Test Run Log
# =====================================================================
def build_test_log_sheet(wb):
    ws = wb.create_sheet("3 - Test Run Log")
    ws.sheet_properties.tabColor = "548235"

    ws.merge_cells("A1:J1")
    ws["A1"].value = "SysId Test Run Log — Record Each Test"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:J2")
    ws["A2"].value = "Fill in one row per test. Yellow cells are input fields."
    ws["A2"].font = SUBTITLE_FONT
    ws["A2"].alignment = Alignment(horizontal="center")

    row = 4
    headers = [
        "Run #",
        "Routine\n(Trans/Steer/Rot)",
        "Test Type\n(QS Fwd / QS Rev /\nDyn Fwd / Dyn Rev)",
        "Battery V\n(Before)",
        "Battery V\n(After)",
        "Duration\nHeld (s)",
        "Robot Behavior\nNormal? (Y/N)",
        "Brownout?\n(Y/N)",
        "Notes / Observations",
        "Operator\nInitials",
    ]
    col_widths = [7, 16, 20, 12, 12, 12, 16, 12, 35, 12]

    for i, (h, w) in enumerate(zip(headers, col_widths), 1):
        ws.cell(row=row, column=i, value=h)
        ws.column_dimensions[get_column_letter(i)].width = w
    style_header_row(ws, row, len(headers))
    ws.row_dimensions[row].height = 45
    row += 1

    # Pre-populate 12 rows (4 tests x 3 routines)
    routines = ["Translation"] * 4 + ["Steer"] * 4 + ["Rotation"] * 4
    test_types = ["QS Forward", "QS Reverse", "Dyn Forward", "Dyn Reverse"] * 3

    for idx, (routine, ttype) in enumerate(zip(routines, test_types), 1):
        ws.cell(row=row, column=1, value=idx).alignment = CENTER
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=2, value=routine).alignment = CENTER
        ws.cell(row=row, column=2).border = THIN_BORDER
        ws.cell(row=row, column=3, value=ttype).alignment = CENTER
        ws.cell(row=row, column=3).border = THIN_BORDER

        # Input cells (yellow)
        for col in range(4, 11):
            style_input_cell(ws, row, col)

        # Add section dividers
        if idx in (1, 5, 9):
            for c in range(1, 3):
                ws.cell(row=row, column=c).fill = SECTION_FILL

        ws.row_dimensions[row].height = 25
        row += 1

    # Extra blank rows for re-runs
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    ws.cell(row=row, column=1, value="EXTRA ROWS (for re-runs if battery swapped or data looked bad)").font = SECTION_FONT
    ws.cell(row=row, column=1).fill = SECTION_FILL
    row += 1

    for extra in range(13, 19):
        ws.cell(row=row, column=1, value=extra).alignment = CENTER
        ws.cell(row=row, column=1).border = THIN_BORDER
        for col in range(2, 11):
            style_input_cell(ws, row, col)
        ws.row_dimensions[row].height = 25
        row += 1


# =====================================================================
# SHEET 4: Results & Gain Update
# =====================================================================
def build_results_sheet(wb):
    ws = wb.create_sheet("4 - Results & Gains")
    ws.sheet_properties.tabColor = "BF8F00"

    ws.merge_cells("A1:H1")
    ws["A1"].value = "SysId Analysis Results & Gain Update Form"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:H2")
    ws["A2"].value = "Record gains from WPILib SysId tool, then update swerve_tuner.py"
    ws["A2"].font = SUBTITLE_FONT
    ws["A2"].alignment = Alignment(horizontal="center")

    # ── Current gains (before) ──
    row = 4
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    ws.cell(row=row, column=1, value="CURRENT GAINS (from swerve_tuner.py — before SysId)").font = SECTION_FONT
    ws.cell(row=row, column=1).fill = SECTION_FILL
    row += 1

    gain_headers = ["Parameter", "kS", "kV", "kA", "kP", "kI", "kD", "Notes"]
    for i, h in enumerate(gain_headers, 1):
        ws.cell(row=row, column=i, value=h)
    style_header_row(ws, row, len(gain_headers))
    row += 1

    current_gains = [
        ("Drive Motors (_drive_gains)", 0.07664325, 0.11412, 0, 0.00442645, 0, 0, "Slot0Configs in swerve_tuner.py"),
        ("Steer Motors (_steer_gains)", 0.1, 1.59, 0, 40, 0, 0.5, "Slot0Configs in swerve_tuner.py"),
    ]
    for label, ks, kv, ka, kp, ki, kd, notes in current_gains:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=1).alignment = LEFT_WRAP
        for col, val in enumerate([ks, kv, ka, kp, ki, kd], 2):
            c = ws.cell(row=row, column=col, value=val)
            c.alignment = CENTER
            c.border = THIN_BORDER
            c.number_format = "0.00000000"
        ws.cell(row=row, column=8, value=notes).alignment = LEFT_WRAP
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=8).border = THIN_BORDER
        row += 1

    # ── New gains (from SysId) ──
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    ws.cell(row=row, column=1, value="NEW GAINS (from WPILib SysId analysis — fill in yellow cells)").font = SECTION_FONT
    ws.cell(row=row, column=1).fill = SECTION_FILL
    row += 1

    for i, h in enumerate(gain_headers, 1):
        ws.cell(row=row, column=i, value=h)
    style_header_row(ws, row, len(gain_headers))
    row += 1

    new_gain_rows = [
        "Drive Motors (NEW)",
        "Steer Motors (NEW)",
        "Rotation / Heading (NEW)",
    ]
    for label in new_gain_rows:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=1).alignment = LEFT_WRAP
        ws.cell(row=row, column=1).border = THIN_BORDER
        for col in range(2, 8):
            style_input_cell(ws, row, col)
        style_input_cell(ws, row, 8)  # notes
        row += 1

    # ── R-squared / fit quality ──
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    ws.cell(row=row, column=1, value="FIT QUALITY (from SysId analysis window)").font = SECTION_FONT
    ws.cell(row=row, column=1).fill = SECTION_FILL
    row += 1

    fit_headers = ["Routine", "R-squared", "Sim Velocity RMSE", "Quality (Good/Fair/Poor)", "Re-run Needed?", "Notes"]
    for i, h in enumerate(fit_headers, 1):
        ws.cell(row=row, column=i, value=h)
    style_header_row(ws, row, len(fit_headers))
    row += 1

    for routine in ["Translation", "Steer", "Rotation"]:
        ws.cell(row=row, column=1, value=routine).font = Font(bold=True)
        ws.cell(row=row, column=1).alignment = CENTER
        ws.cell(row=row, column=1).border = THIN_BORDER
        for col in range(2, 7):
            style_input_cell(ws, row, col)
        row += 1

    # ── Code update checklist ──
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    ws.cell(row=row, column=1, value="CODE UPDATE CHECKLIST").font = SECTION_FONT
    ws.cell(row=row, column=1).fill = SECTION_FILL
    row += 1

    update_steps = [
        ("1", "Open subsystems/drivetrain/swerve_tuner.py"),
        ("2", "Update _drive_gains: kS, kV, kA, kP, kI, kD with NEW Drive Motors values above"),
        ("3", "Update _steer_gains: kS, kV, kA, kP, kI, kD with NEW Steer Motors values above"),
        ("4", "If rotation gains changed: update constants/drive.py kAutoAlign.ROTATION_PID"),
        ("5", "Deploy updated code to roboRIO: python -m robotpy deploy"),
        ("6", "Test drive in Teleop — verify smooth driving, no oscillation"),
        ("7", "Run a known autonomous path — verify tracking accuracy"),
        ("8", "Commit changes to git with descriptive message"),
    ]

    step_headers2 = ["Step", "Action", "Done?"]
    for i, h in enumerate(step_headers2, 1):
        ws.cell(row=row, column=i, value=h)
    style_header_row(ws, row, 3)
    row += 1

    for step, action in update_steps:
        ws.cell(row=row, column=1, value=step).alignment = CENTER
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
        ws.cell(row=row, column=2, value=action).alignment = LEFT_WRAP
        ws.cell(row=row, column=2).border = THIN_BORDER
        style_input_cell(ws, row, 8)
        ws.row_dimensions[row].height = 25
        row += 1

    # ── Sign-off ──
    row += 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    ws.cell(row=row, column=1, value="SIGN-OFF").font = SECTION_FONT
    ws.cell(row=row, column=1).fill = SECTION_FILL
    row += 1

    signoff = [
        "Characterization completed by:",
        "Gains reviewed by:",
        "Code updated & deployed by:",
        "Teleop drive test passed by:",
        "Date completed:",
    ]
    for label in signoff:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        style_input_cell(ws, row, 2)
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 30
    for col_letter in ["B", "C", "D", "E", "F", "G"]:
        ws.column_dimensions[col_letter].width = 14
    ws.column_dimensions["H"].width = 30


# =====================================================================
# SHEET 5: Troubleshooting
# =====================================================================
def build_troubleshooting_sheet(wb):
    ws = wb.create_sheet("5 - Troubleshooting")
    ws.sheet_properties.tabColor = "CC0000"

    ws.merge_cells("A1:D1")
    ws["A1"].value = "SysId Troubleshooting Guide"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center")

    row = 3
    headers = ["Symptom", "Likely Cause", "Fix", "Severity"]
    col_widths = [30, 30, 40, 12]
    for i, (h, w) in enumerate(zip(headers, col_widths), 1):
        ws.cell(row=row, column=i, value=h)
        ws.column_dimensions[get_column_letter(i)].width = w
    style_header_row(ws, row, len(headers))
    row += 1

    issues = [
        (
            "Robot doesn't move when holding A/B/X/Y",
            "Wrong controller port or not in Test mode",
            "Verify Controller 3 is on USB port 2 in DS. Verify DS is in Test mode and Enabled.",
            "HIGH",
        ),
        (
            "Robot browns out during translation test",
            "Battery too low or step voltage too high",
            "Swap to fresh battery (> 12.8V). Translation uses 4V step — if still browning out, reduce in command_swerve_drive.py stepVoltage.",
            "HIGH",
        ),
        (
            "Wheels spin but SysId data is empty/flat",
            "SignalLogger not running",
            "Verify testInit() calls SignalLogger.start(). Check DS console for SignalLogger errors.",
            "HIGH",
        ),
        (
            "HOOT file is 0 bytes or missing",
            "SignalLogger.stop() never called, or disk full",
            "Disable robot before pulling logs (triggers disabledInit). Check free space on roboRIO.",
            "MEDIUM",
        ),
        (
            "SysId tool shows bad R-squared (< 0.9)",
            "Noisy data, battery voltage drop, or too-short test duration",
            "Re-run with fresh battery. Hold QS tests for full 10s. Ensure robot is on blocks (no carpet drag).",
            "MEDIUM",
        ),
        (
            "Steer modules jitter / oscillate after applying new gains",
            "kP too high or kD too low for steer",
            "Reduce steer kP by 25%. Add small kD (start 0.1). Re-test.",
            "HIGH",
        ),
        (
            "Robot drives in circles after gain update",
            "Module gains not symmetric or encoder offsets drifted",
            "Re-check CANcoder offsets in Tuner X. Verify all 4 modules got same gain values.",
            "HIGH",
        ),
        (
            "POV buttons don't switch routine",
            "Wrong controller or binding issue",
            "Verify Controller 3 is connected. Try pressing POV firmly — some controllers have dead zones on D-pad.",
            "LOW",
        ),
        (
            "SysId analysis shows negative kS",
            "Test ran in wrong direction or data is inverted",
            "Re-run the affected tests. Check motor inversion settings in swerve_tuner.py.",
            "MEDIUM",
        ),
        (
            "Robot lurches or overshoots after gain update",
            "kP too aggressive for drive motors",
            "Reduce drive kP by 50%. SysId gains are starting points — tune from there.",
            "MEDIUM",
        ),
    ]

    for symptom, cause, fix, severity in issues:
        ws.cell(row=row, column=1, value=symptom).alignment = LEFT_WRAP
        ws.cell(row=row, column=2, value=cause).alignment = LEFT_WRAP
        ws.cell(row=row, column=3, value=fix).alignment = LEFT_WRAP
        sev_cell = ws.cell(row=row, column=4, value=severity)
        sev_cell.alignment = CENTER
        sev_cell.font = Font(bold=True)
        if severity == "HIGH":
            sev_cell.fill = FAIL_FILL
        elif severity == "MEDIUM":
            sev_cell.fill = WARN_FILL
        else:
            sev_cell.fill = PASS_FILL
        for c in range(1, 5):
            ws.cell(row=row, column=c).border = THIN_BORDER
        ws.row_dimensions[row].height = 50
        row += 1


# =====================================================================
# Main
# =====================================================================
def main():
    wb = Workbook()
    build_cover_sheet(wb)
    build_procedure_sheet(wb)
    build_test_log_sheet(wb)
    build_results_sheet(wb)
    build_troubleshooting_sheet(wb)

    out_path = "docs/SysId_Characterization_Workbook.xlsx"
    wb.save(out_path)
    print(f"Workbook saved to: {out_path}")


if __name__ == "__main__":
    main()
